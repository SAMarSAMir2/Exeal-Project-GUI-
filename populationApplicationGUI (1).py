# The GUI For The Application 0
# project in The University 
# By Samar Samir 



from cProfile import label
from struct import pack
from tkinter import*
from turtle import color
import openpyxl as xl


# creat the main App window 
population_app = Tk()
population_app.config(background='#86A8E7')
population_app.title("Population App")
population_app.geometry("600x600")
the_text = Label(population_app,text="Population App", height=2,width=15, font=("Arial", 40,) ,fg="#007561",bg="#86A8E7")
the_text.place(x=70,y=10)

user = StringVar()

# this function connects the main code eith the GUI
def the_app():
    try:
        user = b.get()
        country = []
        numbers = []
        total = 0
        workbook = xl.load_workbook(f"{user}.xlsx")
        sheet = workbook["Sheet1"]
        for r in range(1,sheet.max_row+1):
            place = sheet.cell(r,1)
            print (f"{place.value}:",end = "")
            country.append(place.value)
            place = sheet.cell(r,2)
            print(place.value)
            total+=place.value
            numbers.append(place.value)
        the_text_1=Label(population_app,text=f"total : {total}\n",fg="#ECE788",bg="#86A8E7",font=("Arial", 15,))
        the_text_1.place(x=50,y=400)
        the_text_2=Label(population_app,text=f"The Hieghest Puolation is  in {country[numbers.index(max(numbers))]} : {max(numbers)} \n",fg="#ECE788",bg="#86A8E7",font=("Arial", 15,))
        the_text_2.place(x=50,y=450)
        the_text_3=Label(population_app,text=f"The minaghest Puolation is  in {country[numbers.index(min(numbers))]}: {min(numbers)} \n",fg="#ECE788",bg="#86A8E7",font=("Arial", 15,))
        the_text_3.place(x=50,y=500)
    except:

        the_text=Label(population_app,text="invalid input" ,fg="#ECE788",bg="#86A8E7",width=30 , height=3,font=("Arial", 15,))
        the_text.place(x=200,y=400)
    

# Great The Enter Button
btn = Button(population_app, text="Enter",font=("Arial", 10) ,width=20 , height=2,fg="#007561",bg="#ECE788", borderwidth=0, command= the_app)
btn.place(x=220,y=270)

# Great The Space Entry
b=StringVar()
text = Entry(textvariable=b ,bg="#ECE788",fg="#494738",font=("Arial", 15) )
text.place(x=180,y=170,width=250 , height=50)


population_app.mainloop()




